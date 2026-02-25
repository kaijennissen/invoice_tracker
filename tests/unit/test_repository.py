"""Tests for invoice_tracker.repository module."""

from pathlib import Path

from openpyxl import load_workbook

from invoice_tracker.repository import (
    ExcelRepository,
    InvoiceRepository,
    create_repository,
)
from invoice_tracker.settings import InvoiceRecord


class TestExcelRepository:
    """Tests for ExcelRepository class."""

    def test_initialize_creates_file(self, tmp_path: Path) -> None:
        """ExcelRepository.initialize creates Excel file with headers."""
        excel_path = tmp_path / "test.xlsx"
        repo = ExcelRepository(excel_path)

        repo.initialize()

        assert excel_path.exists()
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws is not None
        headers = [cell.value for cell in ws[1]]
        assert headers == InvoiceRecord.get_column_headers()
        wb.close()

    def test_save_appends_record(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """ExcelRepository.save appends a record to the Excel file."""
        excel_path = tmp_path / "test.xlsx"
        repo = ExcelRepository(excel_path)
        repo.initialize()

        repo.save(sample_invoice_record)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 2  # 1 header + 1 data row
        assert ws.cell(row=2, column=1).value == sample_invoice_record.party
        wb.close()

    def test_exists_returns_false_for_missing_invoice(self, tmp_path: Path) -> None:
        """ExcelRepository.exists returns False when invoice ID not found."""
        excel_path = tmp_path / "test.xlsx"
        repo = ExcelRepository(excel_path)
        repo.initialize()

        assert repo.exists("NONEXISTENT") is False

    def test_exists_returns_true_for_existing_invoice(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """ExcelRepository.exists returns True when invoice ID is found."""
        excel_path = tmp_path / "test.xlsx"
        repo = ExcelRepository(excel_path)
        repo.initialize()
        repo.save(sample_invoice_record)

        assert repo.exists(sample_invoice_record.invoice_id) is True

    def test_implements_protocol(self, tmp_path: Path) -> None:
        """ExcelRepository satisfies the InvoiceRepository protocol."""
        repo = ExcelRepository(tmp_path / "test.xlsx")
        assert isinstance(repo, InvoiceRepository)


class TestCreateRepository:
    """Tests for create_repository factory function."""

    def test_returns_excel_repository(self, tmp_path: Path) -> None:
        """create_repository returns an ExcelRepository instance."""
        repo = create_repository(tmp_path / "test.xlsx")
        assert isinstance(repo, ExcelRepository)
