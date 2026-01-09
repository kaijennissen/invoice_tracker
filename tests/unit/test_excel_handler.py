"""Tests for invoice_tracker.excel_handler module."""

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from invoice_tracker.excel_handler import append_invoice, init_excel, invoice_exists
from invoice_tracker.settings import InvoiceData, InvoiceRecord


class TestInitExcel:
    """Tests for init_excel function."""

    def test_creates_file_with_headers(self, tmp_path: Path) -> None:
        """init_excel should create a new file with headers."""
        excel_path = tmp_path / "test.xlsx"

        init_excel(excel_path)

        assert excel_path.exists()
        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws is not None
        headers = [cell.value for cell in ws[1]]
        expected_headers = InvoiceRecord.get_column_headers()
        assert headers == expected_headers

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        """init_excel should create parent directories if needed."""
        excel_path = tmp_path / "nested" / "dir" / "test.xlsx"

        init_excel(excel_path)

        assert excel_path.exists()

    def test_skips_existing_file(self, tmp_path: Path) -> None:
        """init_excel should not modify existing file."""
        excel_path = tmp_path / "test.xlsx"

        # Create file first
        init_excel(excel_path)
        original_mtime = excel_path.stat().st_mtime

        # Call again - should not modify
        init_excel(excel_path)

        assert excel_path.stat().st_mtime == original_mtime


class TestAppendInvoice:
    """Tests for append_invoice function."""

    def test_appends_record_to_file(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """append_invoice should add record as new row."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)

        append_invoice(excel_path, sample_invoice_record)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws is not None
        # Row 1 is headers, row 2 should be our data
        data_row = [cell.value for cell in ws[2]]
        assert data_row[0] == sample_invoice_record.party
        assert data_row[1] == sample_invoice_record.invoice_id

    def test_appends_multiple_records(
        self, tmp_path: Path, sample_invoice_data: InvoiceData
    ) -> None:
        """append_invoice should handle multiple records."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)

        record1 = InvoiceRecord(
            **sample_invoice_data.model_dump(),
            source_file="invoice1.png",
            processed_at=datetime.now(),
        )
        record2 = InvoiceRecord(
            party="Other Corp",
            invoice_id="INV-002",
            issue_date=date(2024, 2, 1),
            due_date=date(2024, 3, 1),
            amount=Decimal("500.00"),
            currency="EUR",
            recipient="Jane Doe",
            source_file="invoice2.png",
            processed_at=datetime.now(),
        )

        append_invoice(excel_path, record1)
        append_invoice(excel_path, record2)

        wb = load_workbook(excel_path)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_raises_on_missing_file(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """append_invoice should raise FileNotFoundError for missing file."""
        excel_path = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            append_invoice(excel_path, sample_invoice_record)


class TestInvoiceExists:
    """Tests for invoice_exists function."""

    def test_returns_false_for_missing_file(self, tmp_path: Path) -> None:
        """invoice_exists should return False if file doesn't exist."""
        excel_path = tmp_path / "nonexistent.xlsx"

        assert invoice_exists(excel_path, "INV-001") is False

    def test_returns_false_for_empty_file(self, tmp_path: Path) -> None:
        """invoice_exists should return False for file with only headers."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)

        assert invoice_exists(excel_path, "INV-001") is False

    def test_returns_true_for_existing_invoice(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """invoice_exists should return True if invoice ID is in file."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)
        append_invoice(excel_path, sample_invoice_record)

        assert invoice_exists(excel_path, sample_invoice_record.invoice_id) is True

    def test_returns_false_for_nonexistent_invoice(
        self, tmp_path: Path, sample_invoice_record: InvoiceRecord
    ) -> None:
        """invoice_exists should return False if invoice ID is not in file."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)
        append_invoice(excel_path, sample_invoice_record)

        assert invoice_exists(excel_path, "NONEXISTENT-ID") is False

    def test_finds_invoice_among_many(
        self, tmp_path: Path, sample_invoice_data: InvoiceData
    ) -> None:
        """invoice_exists should find invoice among multiple records."""
        excel_path = tmp_path / "test.xlsx"
        init_excel(excel_path)

        # Add multiple records
        for i in range(5):
            record = InvoiceRecord(
                party=f"Company {i}",
                invoice_id=f"INV-{i:03d}",
                issue_date=date(2024, 1, i + 1),
                due_date=date(2024, 2, i + 1),
                amount=Decimal(str(100 * (i + 1))),
                currency="EUR",
                recipient=f"Recipient {i}",
                source_file=f"invoice_{i}.png",
                processed_at=datetime.now(),
            )
            append_invoice(excel_path, record)

        # Check middle record exists
        assert invoice_exists(excel_path, "INV-002") is True
        # Check non-existent
        assert invoice_exists(excel_path, "INV-999") is False
