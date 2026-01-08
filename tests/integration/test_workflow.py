"""Integration tests for the invoice processing workflow.

These tests verify the end-to-end workflow with mocked Ollama responses.
They are marked with @pytest.mark.integration and may be slower than unit tests.
"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from invoice_tracker.processor import process_batch
from invoice_tracker.settings import InvoiceData, Settings


@pytest.fixture
def integration_settings(tmp_path: Path) -> Settings:
    """Create settings for integration testing.

    Parameters
    ----------
    tmp_path : Path
        Pytest temporary path fixture.

    Returns
    -------
    Settings
        Fully configured settings for integration testing.
    """
    incoming = tmp_path / "invoices" / "incoming"
    processed = tmp_path / "invoices" / "processed"
    failed = tmp_path / "invoices" / "failed"
    excel_file = tmp_path / "data" / "tracker.xlsx"

    incoming.mkdir(parents=True)
    processed.mkdir(parents=True)
    failed.mkdir(parents=True)
    excel_file.parent.mkdir(parents=True)

    return Settings(
        _cli_parse_args=False,
        incoming_dir=incoming,
        processed_dir=processed,
        failed_dir=failed,
        excel_file=excel_file,
    )


@pytest.mark.integration
class TestFullWorkflow:
    """Integration tests for the full invoice processing workflow."""

    def test_process_batch_creates_excel_and_moves_files(
        self, integration_settings: Settings
    ) -> None:
        """Full workflow: scan, extract, persist to Excel, move files."""
        # Create test invoice files
        (integration_settings.incoming_dir / "invoice_001.png").touch()
        (integration_settings.incoming_dir / "invoice_002.jpg").touch()

        # Mock Ollama extraction with different invoice data
        mock_invoices = [
            InvoiceData(
                party="Company A",
                invoice_id="INV-2024-001",
                issue_date=date(2024, 1, 15),
                due_date=date(2024, 2, 15),
                amount=Decimal("1000.00"),
                currency="EUR",
                recipient="John Doe",
            ),
            InvoiceData(
                party="Company B",
                invoice_id="INV-2024-002",
                issue_date=date(2024, 1, 20),
                due_date=date(2024, 2, 20),
                amount=Decimal("2500.50"),
                currency="USD",
                recipient="Jane Smith",
            ),
        ]

        with patch(
            "invoice_tracker.processor.extract_invoice"
        ) as mock_extract:
            mock_extract.side_effect = mock_invoices

            results = process_batch(integration_settings)

        # Verify results
        assert len(results) == 2
        assert all(r.success for r in results)

        # Verify files moved to processed
        assert not (integration_settings.incoming_dir / "invoice_001.png").exists()
        assert not (integration_settings.incoming_dir / "invoice_002.jpg").exists()
        assert (integration_settings.processed_dir / "invoice_001.png").exists()
        assert (integration_settings.processed_dir / "invoice_002.jpg").exists()

        # Verify Excel file created with data
        assert integration_settings.excel_file.exists()
        wb = load_workbook(integration_settings.excel_file)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 3  # Header + 2 data rows

        # Verify first data row
        assert ws.cell(2, 1).value == "Company A"  # party
        assert ws.cell(2, 2).value == "INV-2024-001"  # invoice_id

    def test_failed_extraction_moves_to_failed_dir(
        self, integration_settings: Settings
    ) -> None:
        """Failed extractions should move files to failed directory."""
        (integration_settings.incoming_dir / "bad_invoice.png").touch()

        with patch(
            "invoice_tracker.processor.extract_invoice"
        ) as mock_extract:
            from invoice_tracker.settings import ExtractionError

            mock_extract.side_effect = ExtractionError("Invalid image")

            results = process_batch(integration_settings)

        assert len(results) == 1
        assert not results[0].success
        assert "Extraction failed" in results[0].error  # type: ignore[operator]

        # Verify file moved to failed
        assert not (integration_settings.incoming_dir / "bad_invoice.png").exists()
        assert (integration_settings.failed_dir / "bad_invoice.png").exists()

    def test_dry_run_does_not_modify_filesystem(
        self, integration_settings: Settings
    ) -> None:
        """Dry run should extract but not persist or move files."""
        integration_settings.dry_run = True
        (integration_settings.incoming_dir / "invoice.png").touch()

        mock_invoice = InvoiceData(
            party="Test Company",
            invoice_id="INV-TEST-001",
            issue_date=date(2024, 1, 1),
            due_date=date(2024, 2, 1),
            amount=Decimal("500.00"),
            currency="EUR",
            recipient="Test User",
        )

        with patch(
            "invoice_tracker.processor.extract_invoice"
        ) as mock_extract:
            mock_extract.return_value = mock_invoice

            results = process_batch(integration_settings)

        assert len(results) == 1
        assert results[0].success

        # Verify file NOT moved
        assert (integration_settings.incoming_dir / "invoice.png").exists()
        assert not (integration_settings.processed_dir / "invoice.png").exists()

        # Verify Excel NOT created
        assert not integration_settings.excel_file.exists()
