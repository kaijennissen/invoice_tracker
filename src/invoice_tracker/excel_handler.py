"""Excel file handler for invoice persistence.

This module provides functions to read and write invoice data to Excel files.
It is part of the persistence/IO layer and should be called from orchestration
functions, not directly from business logic.

Uses openpyxl for Excel file operations and derives headers from the
InvoiceRecord model.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from invoice_tracker.retry import RetryConfig, with_retry
from invoice_tracker.settings import InvoiceRecord

_EXCEL_RETRY = RetryConfig(max_retries=2, initial_backoff=0.5, catch=(PermissionError,))


def _get_invoice_id_column_index() -> int:
    """Get 1-indexed column position of invoice_id field.

    Returns
    -------
    int
        The 1-indexed column number for the invoice_id field.
    """
    field_names = list(InvoiceRecord.model_fields.keys())
    return field_names.index("invoice_id") + 1


def init_excel(path: Path) -> None:
    """Create workbook with headers derived from InvoiceRecord model.

    If the file already exists, this function does nothing.
    Creates parent directories if they don't exist.

    Parameters
    ----------
    path : Path
        Path to the Excel file to create.
    """
    if path.exists():
        return

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.append(InvoiceRecord.get_column_headers())
    wb.save(path)


@with_retry(_EXCEL_RETRY)
def _save_workbook(path: Path, values: list[Any]) -> None:
    """Load workbook, append values, and save.

    Parameters
    ----------
    path : Path
        Path to the Excel file.
    values : list[Any]
        Row values to append.

    Raises
    ------
    PermissionError
        If the file is locked.
    ValueError
        If the workbook has no active worksheet.
    """
    wb = load_workbook(path)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active worksheet")
        ws.append(values)
        wb.save(path)
    finally:
        wb.close()


def append_invoice(path: Path, record: InvoiceRecord) -> None:
    """Append invoice record to Excel file.

    Uses retry logic with exponential backoff for file locking issues.

    Parameters
    ----------
    path : Path
        Path to the Excel file.
    record : InvoiceRecord
        Invoice record to append.

    Raises
    ------
    PermissionError
        If the file is locked after all retries.
    FileNotFoundError
        If the Excel file doesn't exist.
    ValueError
        If the workbook has no active worksheet.
    """
    values = list(record.model_dump().values())
    _save_workbook(path, values)


def invoice_exists(path: Path, invoice_id: str) -> bool:
    """Check if an invoice with the given ID already exists in the file.

    Parameters
    ----------
    path : Path
        Path to the Excel file.
    invoice_id : str
        Invoice ID to check for.

    Returns
    -------
    bool
        True if the invoice ID exists in the file, False otherwise.
    """
    if not path.exists():
        return False

    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        if ws is None:
            return False

        # Derive invoice_id column index from model to avoid hardcoding
        invoice_id_col = _get_invoice_id_column_index()

        # Skip header row, check all data rows
        for row in ws.iter_rows(
            min_row=2, min_col=invoice_id_col, max_col=invoice_id_col
        ):
            cell = row[0]
            if cell.value == invoice_id:
                return True

        return False
    finally:
        wb.close()


__all__ = [
    "init_excel",
    "append_invoice",
    "invoice_exists",
]
